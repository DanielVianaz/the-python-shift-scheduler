import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def gerar_excel(escala, nome_arquivo="escala.xlsx"):

    wb = Workbook()
    ws = wb.active
    ws.title = "Escala"

    operadores = ["Sofia", "Jorge", "Viviane", "Daniel"]

    # - ESTILOS

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    total_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

    bold = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # - CABEÇALHO

    cabecalho = [
        "Dia",
        "Sofia", "Horas",
        "Jorge", "Horas",
        "Viviane", "Horas",
        "Daniel", "Horas"
    ]

    ws.append(cabecalho)

    for col in range(1, len(cabecalho) + 1):

        cell = ws.cell(row=1, column=col)

        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # - CONTADOR HORAS

    horas = {op: 0 for op in operadores}

    # -ESCALA

    for i, dia in enumerate(escala, start=1):

        linha = [i]

        turnos = {op: "" for op in operadores}
        horas_dia = {op: "" for op in operadores}

        # abertura
        if "abertura" in dia:

            op = dia["abertura"]

            horario = dia.get("abertura_turno", "09:30-21:30")
            h = dia.get("abertura_horas", 12)

            turnos[op] = horario
            horas_dia[op] = h
            horas[op] += h

        # fecho
        if "fecho" in dia:

            op = dia["fecho"]

            horario = dia.get("fecho_turno", "11:00-23:00")
            h = dia.get("fecho_horas", 12)

            turnos[op] = horario
            horas_dia[op] = h
            horas[op] += h

        for op in operadores:

            linha.append(turnos[op])
            linha.append(horas_dia[op])

        ws.append(linha)

    # - BORDAS NA TABELA

    max_row = ws.max_row

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            cell.alignment = center

    # - LINHA TOTAL

    linha_total = ws.max_row + 1

    ws.cell(row=linha_total, column=1).value = "TOTAL"
    ws.cell(row=linha_total, column=1).font = bold
    ws.cell(row=linha_total, column=1).fill = total_fill
    ws.cell(row=linha_total, column=1).alignment = center
    ws.cell(row=linha_total, column=1).border = border

    col = 2

    for op in operadores:

        ws.cell(row=linha_total, column=col).value = ""
        ws.cell(row=linha_total, column=col).border = border

        total_cell = ws.cell(row=linha_total, column=col + 1)
        total_cell.value = horas[op]
        total_cell.font = bold
        total_cell.fill = total_fill
        total_cell.alignment = center
        total_cell.border = border

        col += 2

    # - LARGURA DAS COLUNAS

    ws.column_dimensions["A"].width = 6

    for col in ["B", "D", "F", "H"]:
        ws.column_dimensions[col].width = 18

    for col in ["C", "E", "G", "I"]:
        ws.column_dimensions[col].width = 8

    # - SALVAR

    wb.save(nome_arquivo)

    print("Excel gerado:", nome_arquivo)

    try:
        os.startfile(nome_arquivo)
    except:
        pass